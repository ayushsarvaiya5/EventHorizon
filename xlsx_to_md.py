"""
xlsx_to_md.py
Converts every sheet in an .xlsx file into a GitHub-flavoured Markdown file.
Each sheet becomes a ## section with a pipe table.

Usage:
    python xlsx_to_md.py
    python xlsx_to_md.py --input "path/to/file.xlsx" --output "path/to/output.md"
"""

import argparse
import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("[ERROR] openpyxl not found. Run: pip install openpyxl tabulate")

try:
    from tabulate import tabulate
except ImportError:
    sys.exit("[ERROR] tabulate not found. Run: pip install openpyxl tabulate")


DEFAULT_INPUT = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "EVENT CONFIGURATION - Access Control and TIme Attendance.xlsx",
)
DEFAULT_OUTPUT = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "EVENT CONFIGURATION - Access Control and TIme Attendance.md",
)


def cell_value(cell) -> str:
    """Return a clean string from an openpyxl cell, handling None and newlines."""
    if cell.value is None:
        return ""
    # Collapse internal newlines so table rows stay on one line
    return str(cell.value).replace("\n", " ").replace("\r", " ").strip()


def sheet_to_md(ws) -> str:
    """Convert a single worksheet to a Markdown pipe-table string."""
    rows = list(ws.iter_rows())
    if not rows:
        return "_Sheet is empty._\n"

    # First row = headers
    headers = [cell_value(c) for c in rows[0]]
    # Use generic column names if header row is completely blank
    if all(h == "" for h in headers):
        headers = [f"Column {i + 1}" for i in range(len(rows[0]))]

    data_rows = []
    for row in rows[1:]:
        values = [cell_value(c) for c in row]
        # Skip rows that are entirely empty
        if any(v != "" for v in values):
            data_rows.append(values)

    if not data_rows:
        return "_Sheet has headers but no data rows._\n"

    # Ensure every data row has the same column count as headers
    col_count = len(headers)
    padded = [r + [""] * (col_count - len(r)) if len(r) < col_count else r[:col_count]
              for r in data_rows]

    md_table = tabulate(padded, headers=headers, tablefmt="pipe", missingval="")
    return md_table + "\n"


def convert(input_path: str, output_path: str) -> None:
    if not os.path.isfile(input_path):
        sys.exit(f"[ERROR] Input file not found: {input_path}")

    print(f"Reading: {input_path}")
    wb = openpyxl.load_workbook(input_path, data_only=True)

    sections = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  Processing sheet: '{sheet_name}' ({ws.max_row} rows × {ws.max_column} cols)")
        md_block = f"## {sheet_name}\n\n{sheet_to_md(ws)}"
        sections.append(md_block)

    title = os.path.splitext(os.path.basename(input_path))[0]
    content = f"# {title}\n\n" + "\n\n---\n\n".join(sections) + "\n"

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(content)

    print(f"\nOutput written to: {output_path}")
    print(f"Total sheets converted: {len(wb.sheetnames)}")


def main():
    parser = argparse.ArgumentParser(description="Convert .xlsx to Markdown tables")
    parser.add_argument("--input", default=DEFAULT_INPUT, help="Path to input .xlsx file")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Path to output .md file")
    args = parser.parse_args()
    convert(args.input, args.output)


if __name__ == "__main__":
    main()
